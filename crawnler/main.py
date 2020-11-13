# encoding:utf-8


import util.util as util
import open_url
import anylize_searchresult
import anylize_detail
import output
import os
import sys
import configparser

from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.parse


def search_keyword(keyword):
    key_decode = urllib.parse.quote(keyword)
    url = util.search_Url + key_decode
    str_web = open_url.open_url(url)

    try:
        company_page_total = anylize_searchresult.get_total_number(keyword, str_web)
        print("爬取总页数：", company_page_total, sep=" ")
        company_list = anylize_searchresult.get_company(keyword, str_web)
        # 可修改爬取页数，注意range(2, company_page_rotal+1) 第一个参数是从第几页开始爬取，一定大于等于2
        for i in range(2, company_page_total+1):
            url_page = url + "&p=" + i.__str__() + "&"
            str_web = open_url.open_url(url_page)

            company_list_more = anylize_searchresult.get_company(keyword, str_web)
            for item_result in company_list_more[1]:
                company_list[1].append(item_result)
    except:
        q_str1 = "window.location.href='http://www.qichacha.com/index_verify"
        q_str2 = "window.location.href='https://www.qichacha.com/index_verify"
        if not ((q_str1 in str_web) or (q_str2 in str_web)):
            print("无法解析页面的关键词：", keyword, sep=" ")
            output.output_fail([keyword])
            return
        else:
            sys.exit(0)

    if len(company_list) == 0:
        output.output([keyword, "没有结果"])
    else:
        for item, company_name in company_list[1]:
            if item[0] == "/":
                item = util.qichacha_Url + item

            str_web = open_url.open_url(item)

            try:
                company_detail = anylize_detail.get_detail(keyword, company_name, str_web)
                pass
            except:
                q_str = "window.location.href='http://www.qichacha.com/index_verify"
                q_str2 = "window.location.href='https://www.qichacha.com/index_verify"
                if not ((q_str in str_web) or (q_str2 in str_web)):
                    print("无法解析页面的关键词：", keyword, company_name, sep=" ")
                    output.output_fail([keyword, company_name])
                else:
                    sys.exit(0)
            else:
                # output.output(company_detail)
                try:
                    output.output(company_detail)
                except:
                    print("\n公司详细信息内容包含 excel 不支持字符，写入失败")
                    print("公司名称：", company_name, sep=" ")
                    print("\n")
                    output.output_fail([keyword, company_name])

def copy_file (inputFile, outputFile, encoding):
    fin = open(inputFile, 'r', encoding=encoding) #以读的方式打开文件
    fout = open(outputFile, 'w', encoding=encoding) #以写得方式打开文件
    for eachLiine in fin.readlines(): #读取文件的每一行
        line = eachLiine.strip().replace("\ufeff", "") #去除每行的首位空格
        fout.write(line + '\n')
    fin.close()
    fout.close()


def get_conf(str):
    copy_file(util.config_file, util.config_encode_file, "utf-8")

    cf = configparser.ConfigParser()
    cf.read(util.config_encode_file, encoding="utf-8")
    return cf.get("path", str)


if __name__ == '__main__':
    if not os.path.exists(util.result_file):
        os.mkdir(util.result_file)
    if not os.path.exists(util.save_data_name):
        wb = Workbook()
        sheet = wb.get_active_sheet()
        sheet.title = "公司信息表"
        sheet.append(util.detail_name)
        wb.save(util.save_data_name)

    if not os.path.exists(util.fail_data_name):
        wb = Workbook()
        sheet = wb.get_active_sheet()
        sheet.title = "搜索失败关键词"
        sheet.append(["关键词"])
        wb.save(util.fail_data_name)


    wb = load_workbook(get_conf("file_test_name"))
    sheet = wb.get_sheet_by_name(get_conf("file_sheet_name"))
    # 可以修改第几行开始读数据，[:]包含第一个数，不包含第二个数，第一行从0开始标号
    for item in sheet["A"][:]:
        keyword = item.value
        if keyword == None:
            break
        keyword = str(keyword)
        print("\n", keyword)
        search_keyword(keyword)
        
    os.system('pause') 
