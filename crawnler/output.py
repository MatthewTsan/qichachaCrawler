# encoding:utf-8

import util.util as util
import csv
from openpyxl import load_workbook

def output_fail(list):
    wb = load_workbook(util.fail_data_name)
    sheet = wb.get_active_sheet()
    sheet.append(list)
    wb.save(util.fail_data_name)

def output(list):
    # dict = {}
    # print(len(list), len(util.detail_name), sep=" ")
    # for x, item in enumerate(util.detail_name):
    #     dict[item] = list[x]
    # print(dict)
    # # dataframe = pd.DataFrame(dict)
    #
    # # dataframe.to_csv("test.csv", index=False, sep=",")


    wb = load_workbook(util.save_data_name)
    sheet = wb.get_active_sheet()
    if len(list) == 2:
        sheet.append(list)
    else:
        for i in range(len(list[util.change_list_pos])):
            result = []
            for num, item in enumerate(list):
                if num == util.change_list_pos:
                    for data in list[util.change_list_pos][i]:
                        result.append(data)
                else:
                    result.append(item)
            sheet.append(result)
    wb.save(util.save_data_name)